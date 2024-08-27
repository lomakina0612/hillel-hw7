import json
import requests

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Font, Alignment
from datetime import datetime


rates =[]
for year in range(2021, 2024):
   for month in range(1,13):
      r = requests.get(f'https://api.privatbank.ua/p24api/exchange_rates?json&date=01.{month:02}.{year}')
      
      if r.status_code == 200:
         response = json.loads(r.text)
         rate_eur = [item for item in response['exchangeRate'] if item['currency'] == 'EUR']

      print(f'01.{month:02}.{year} 1 EUR = {rate_eur[0]['saleRate']:.2f} UAH') 
      rates.append((f'01.{month:02}.{year}',rate_eur[0]['saleRate']))
print(rates)


wb = Workbook()
ws = wb.active
ws.title = 'EUR exchange rate'
ws["A1"] = 'Date'
ws["B1"] = 'EUR'

for row in ws['A1:B1']:
    for cell in row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

for idx, rate in enumerate(rates, start=2):
    date_object = datetime.strptime(rate[0], '%d.%m.%Y')
    ws[f'A{idx}'] = date_object
    ws[f'B{idx}'] = rate[1]

for row in ws.iter_rows(min_row=2, max_col=2):
    row[0].number_format = 'dd.mm.yy'
    row[1].number_format = '0.00'

ws.column_dimensions['A'].width = 11
       
c1= LineChart()
c1.title = "EUR to UAH exchange rate"
c1.style = 12

data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(rates)+1)
c1.add_data(data, titles_from_data=True)
dates = Reference(ws, min_col=1, min_row=2, max_row=len(rates)+1)
c1.set_categories(dates)

s1 = c1.series[0]
s1.marker.symbol = "circle"
s1.marker.graphicalProperties.solidFill = "FFFFFF" # Marker filling
s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline
s1.graphicalProperties.line.solidFill = "FF0000" # Line
s1.graphicalProperties.line.width = 40000 # width in EMUs

c1.y_axis.title = 'Rate'
c1.y_axis.crossAx = 500
c1.x_axis = DateAxis(crossAx=100)
c1.x_axis.number_format = "dd.mm.yyyy"
c1.x_axis.majorTimeUnit = "months"
c1.x_axis.title = 'Date'

c1.width = 25  
c1.height = 10  

ws.add_chart(c1, "D1")
wb.save("rates.xlsx")
